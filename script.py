import openpyxl
from pathlib import Path
import re
import ast
import sys
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.cell import get_column_letter, range_boundaries
import tkinter as tk
from tkinter import filedialog

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# КОНСТАНТИ для назв товарів-остачі
OSTACHA_NAMES = {
    '2210': 'Футболки з нанесенням логотипу',
    '2240': 'Перевезення',
    '2250': 'Добові'
}

# Колір для виділення остачі (блакитний)
OSTACHA_COLOR = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")


def select_file(title, default_filename):
    """Вибір файлу через діалогове вікно"""
    root = tk.Tk()
    root.withdraw()  # Ховаємо головне вікно
    root.attributes('-topmost', True)  # Вікно поверх інших

    file_path = filedialog.askopenfilename(
        title=title,
        initialfile=default_filename,
        filetypes=[
            ("Excel files", "*.xlsx *.xls *.xlsm"),
            ("All files", "*.*")
        ]
    )

    root.destroy()

    if file_path:
        return Path(file_path)
    else:
        return None


def is_rozrahunok_candidate(file_path):
    """Перевіряє, чи Excel-файл схожий на Розрахунок, а не на Кошторис/Зміни."""
    name_lower = file_path.name.lower()
    if file_path.name.startswith('~$'):
        return False
    if any(skip_word in name_lower for skip_word in ['кошторис', 'зміни']):
        return False
    if file_path.suffix.lower() not in ['.xlsx', '.xlsm']:
        return False

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active

        if ws.max_column < 22:
            wb.close()
            return False

        for row in range(1, min(ws.max_row, 100) + 1):
            a_val = ws[f'A{row}'].value
            b_val = ws[f'B{row}'].value
            if isinstance(a_val, (int, float)) and isinstance(b_val, str):
                money_cells = [ws[f'{col}{row}'].value for col in ['L', 'M', 'N', 'T', 'U', 'V']]
                if any(value is not None for value in money_cells):
                    wb.close()
                    return True

        wb.close()
    except Exception:
        return False

    return False


def find_rozrahunok_file(default_filename):
    """Шукає файл Розрахунку в поточній папці."""
    default_file = Path(default_filename)
    if default_file.exists():
        return default_file

    candidates = [
        file_path for file_path in Path('.').glob('*.xls*')
        if is_rozrahunok_candidate(file_path)
    ]

    if not candidates:
        return None

    def candidate_score(file_path):
        name_lower = file_path.name.lower()
        score = 0
        if 'розрах' in name_lower or 'rozrah' in name_lower:
            score += 10
        if 'приват' in name_lower:
            score += 2
        return score

    candidates.sort(key=lambda path: (candidate_score(path), path.stat().st_mtime), reverse=True)
    return candidates[0]


def is_purchased_item(item_name):
    """Перевіряє чи товар є купленим (має закінчення (1812))"""
    if not item_name:
        return False
    return bool(re.search(r'\(1812\)\s*$', str(item_name)))


def calculate_formula_value(formula_str, row_num, ws):
    """Вичисляє значення формули типа =D2*F2 або =G3+G4"""
    if not formula_str or not isinstance(formula_str, str):
        return formula_str

    if not formula_str.startswith('='):
        return formula_str

    formula = formula_str[1:]

    # Підставляємо значення для поточного рядка
    for r in range(2, 15):
        for col_letter in ['D', 'E', 'F', 'G']:
            cell_val = ws[f'{col_letter}{r}'].value
            if isinstance(cell_val, str) and cell_val.startswith('='):
                try:
                    d_val = ws[f'D{r}'].value or 0
                    e_val = ws[f'E{r}'].value or 0
                    f_val = ws[f'F{r}'].value or 0
                    nested_formula = cell_val[1:]
                    nested_result = eval(nested_formula.replace(f'D{r}', str(d_val))
                                         .replace(f'E{r}', str(e_val))
                                         .replace(f'F{r}', str(f_val)))
                    formula = formula.replace(f'{col_letter}{r}', str(nested_result))
                except:
                    formula = formula.replace(f'{col_letter}{r}', '0')
            else:
                val = cell_val or 0
                formula = formula.replace(f'{col_letter}{r}', str(val))

    try:
        result = eval(formula)
        return result
    except:
        return 0


def parse_number(value):
    """Перетворює число з Excel/тексту в float або повертає None."""
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        normalized = value.strip().replace(' ', '').replace(',', '.')
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


def safe_eval_arithmetic(expression):
    """Безпечно рахує прості арифметичні формули."""
    operators = {
        ast.Add: lambda left, right: left + right,
        ast.Sub: lambda left, right: left - right,
        ast.Mult: lambda left, right: left * right,
        ast.Div: lambda left, right: left / right if right != 0 else 0,
        ast.USub: lambda value: -value,
        ast.UAdd: lambda value: value,
    }

    def eval_node(node):
        if isinstance(node, ast.Expression):
            return eval_node(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return node.value
        if isinstance(node, ast.Num):
            return node.n
        if isinstance(node, ast.BinOp) and type(node.op) in operators:
            return operators[type(node.op)](eval_node(node.left), eval_node(node.right))
        if isinstance(node, ast.UnaryOp) and type(node.op) in operators:
            return operators[type(node.op)](eval_node(node.operand))
        raise ValueError("Unsupported formula")

    try:
        return eval_node(ast.parse(expression, mode='eval'))
    except Exception:
        return None


def expand_sum_ranges(formula, ws_values, ws_formulas, seen):
    """Підтримка простих SUM(A1:A3) у формулах Excel."""
    def sum_replacer(match):
        total = 0
        args = re.split(r'[;,]', match.group(1))
        for arg in args:
            arg = arg.strip().replace('$', '')
            if ':' in arg:
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(arg)
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            address = f'{get_column_letter(col)}{row}'
                            total += parse_number(get_cell_value(ws_values, address, ws_formulas, seen)) or 0
                except Exception:
                    return '0'
            else:
                total += parse_number(get_cell_value(ws_values, arg, ws_formulas, seen)) or 0
        return str(total)

    return re.sub(r'SUM\(([^()]*)\)', sum_replacer, formula, flags=re.IGNORECASE)


def calculate_excel_formula(formula, ws_values, ws_formulas, seen=None):
    """Рахує прості Excel-формули з посиланнями на комірки."""
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return parse_number(formula)

    if seen is None:
        seen = set()

    expression = formula[1:].replace('$', '')
    expression = expand_sum_ranges(expression, ws_values, ws_formulas, seen)

    def ref_replacer(match):
        address = match.group(0).replace('$', '')
        return str(parse_number(get_cell_value(ws_values, address, ws_formulas, seen)) or 0)

    expression = re.sub(r'\b[A-Z]{1,3}\d+\b', ref_replacer, expression)
    expression = expression.replace(',', '.')

    if not re.fullmatch(r'[0-9.\s+\-*/()]+', expression):
        return None

    return safe_eval_arithmetic(expression)


def get_cell_value(ws, cell_address, ws_formulas=None, seen=None):
    """Отримує вичислене значення комірки, навіть якщо Excel не зберіг кеш формули."""
    if seen is None:
        seen = set()

    cell_address = str(cell_address).replace('$', '')
    if cell_address in seen:
        return 0

    cell = ws[cell_address]
    numeric_value = parse_number(cell.value)
    if numeric_value is not None:
        return numeric_value

    if cell.value is not None and not isinstance(cell.value, str):
        return cell.value

    formula_value = None
    if ws_formulas is not None:
        formula_value = ws_formulas[cell_address].value
    elif isinstance(cell.value, str) and cell.value.startswith('='):
        formula_value = cell.value

    if isinstance(formula_value, str) and formula_value.startswith('='):
        seen.add(cell_address)
        calculated_value = calculate_excel_formula(formula_value, ws, ws_formulas or ws, seen)
        seen.discard(cell_address)
        if calculated_value is not None:
            return calculated_value

    return cell.value


def parse_formula_references(formula_str):
    """Витягує всі посилання на комірки з формули"""
    if not formula_str or not isinstance(formula_str, str):
        return []
    if formula_str.startswith('='):
        formula_str = formula_str[1:]
    pattern = r'[A-Z]+\d+'
    matches = re.findall(pattern, formula_str)
    return matches


def get_kekv_mapping_for_zahid(rozrahunok_file, zahid_start_row):
    """Визначає які рядки належать до яких КЕКВ на основі формул"""
    kekv_mapping = {}

    wb_formulas = openpyxl.load_workbook(rozrahunok_file, data_only=False)
    ws_formulas = wb_formulas.active

    kekv_columns = {'L': '2210', 'M': '2240', 'N': '2250'}

    for col_letter, kekv_code in kekv_columns.items():
        cell_address = f'{col_letter}{zahid_start_row}'
        cell = ws_formulas[cell_address]

        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            references = parse_formula_references(cell.value)
            for ref in references:
                row_match = re.search(r'\d+', ref)
                if row_match:
                    row_num = int(row_match.group())
                    if row_num not in kekv_mapping:
                        kekv_mapping[row_num] = []
                    kekv_mapping[row_num].append(kekv_code)

    wb_formulas.close()
    return kekv_mapping


def find_zahid_in_rozrahunok(ws_rozrahunok, zahid_number):
    """Знаходить захід в Розрахунку за номером"""
    for row in range(2, ws_rozrahunok.max_row + 1):
        a_val = ws_rozrahunok[f'A{row}'].value
        if a_val is not None and isinstance(a_val, (int, float)) and int(a_val) == zahid_number:
            b_val = ws_rozrahunok[f'B{row}'].value  # Назва
            c_val = ws_rozrahunok[f'C{row}'].value  # Термін

            # Визначаємо межі заходу
            end_row = row
            for next_row in range(row + 1, ws_rozrahunok.max_row + 1):
                next_a_val = ws_rozrahunok[f'A{next_row}'].value
                if next_a_val is not None and isinstance(next_a_val, (int, float)):
                    end_row = next_row - 1
                    break
            else:
                end_row = ws_rozrahunok.max_row

            return {
                'number': int(a_val),
                'name': b_val,
                'termin': c_val,
                'start_row': row,
                'end_row': end_row
            }
    return None


def get_zalushky_for_zahid(ws_rozrahunok, zahid_start_row, ws_rozrahunok_formulas=None):
    """Отримує залишки по КЕКВ для заходу з колонок T, U, V"""
    return {
        '2210': get_cell_value(ws_rozrahunok, f'T{zahid_start_row}', ws_rozrahunok_formulas) or 0,
        '2240': get_cell_value(ws_rozrahunok, f'U{zahid_start_row}', ws_rozrahunok_formulas) or 0,
        '2250': get_cell_value(ws_rozrahunok, f'V{zahid_start_row}', ws_rozrahunok_formulas) or 0
    }


def get_koshtorys_needs(koshtorys_data):
    """Розраховує потребу по кожному КЕКВ з Кошторису (БЕЗ куплених товарів (1812))"""
    _, _, items_koshtorys, kekv_data, kekv_purchased, _ = koshtorys_data

    # Рахуємо потребу БЕЗ товарів (1812)
    needs = {'2210': 0, '2240': 0, '2250': 0}

    for item in items_koshtorys:
        if not is_purchased_item(item['name']):
            kekv = item['kekv']
            if kekv in needs:
                needs[kekv] += item['suma']

    return needs


def get_remaining_needs(koshtorys_needs, already_taken):
    """Рахує, скільки ще треба взяти по кожному КЕКВ."""
    return {
        kekv: max(0, koshtorys_needs.get(kekv, 0) - already_taken.get(kekv, 0))
        for kekv in ['2210', '2240', '2250']
    }


def print_big_deficit_warning(deficits):
    """Друкує максимально помітне попередження про нестачу коштів."""
    print("\n" + "!" * 80)
    print("!!! УВАГА !!! НЕ ВИСТАЧАЄ ГРОШЕЙ !!! УВАГА !!!")
    print("!" * 80)
    print("По цих КЕКВ коштів менше, ніж потрібно за Кошторисом:")
    for kekv in ['2210', '2240', '2250']:
        deficit = deficits.get(kekv, 0)
        if deficit > 0:
            print(f"   КЕКВ {kekv}: НЕ ВИСТАЧАЄ {deficit:.2f} грн")
    print("!" * 80)
    print("Зміни НЕ будуть створені без окремого підтвердження.")
    print("!" * 80)


def find_closest_match(target, values, tolerance=100):
    """Знаходить найближче значення до цільового (шукає пару для підрахунку залишку)"""
    best_match = None
    min_diff = float('inf')

    for val in values:
        if val <= target:
            diff = abs(target - val)
            if diff <= tolerance and diff < min_diff:
                min_diff = diff
                best_match = val

    return best_match


def get_tovary_by_kekv(ws_rozrahunok, zahid, zalushky, kekv_mapping, allowed_kekv, rozrahunok_file):
    """Отримує список товарів з залишками (порівнює формули L/M/N з P/Q/R)"""
    tovary_by_kekv = {'2210': [], '2240': [], '2250': []}

    wb_formulas = openpyxl.load_workbook(rozrahunok_file, data_only=False)
    ws_formulas = wb_formulas.active

    kekv_columns = {
        '2210': {'plan': 'L', 'used': 'P', 'zalushok': 'T'},
        '2240': {'plan': 'M', 'used': 'Q', 'zalushok': 'U'},
        '2250': {'plan': 'N', 'used': 'R', 'zalushok': 'V'}
    }

    for kekv in ['2210', '2240', '2250']:
        if kekv not in allowed_kekv:
            continue

        cols = kekv_columns[kekv]
        real_zalushok = zalushky.get(kekv, 0)
        if real_zalushok <= 0:
            continue

        print(f"\n  Аналіз КЕКВ {kekv}:")
        print(f"    Реальний залишок з {cols['zalushok']}: {real_zalushok}")

        all_plan_items = {}
        plan_cell = ws_formulas[f'{cols["plan"]}{zahid["start_row"]}']

        if plan_cell.value and isinstance(plan_cell.value, str) and plan_cell.value.startswith('='):
            plan_refs = parse_formula_references(plan_cell.value)
            print(f"    План {cols['plan']}: {plan_cell.value}")

            for ref in plan_refs:
                row_match = re.search(r'\d+', ref)
                if row_match:
                    row_num = int(row_match.group())
                    h_val = ws_rozrahunok[f'H{row_num}'].value
                    k_val = get_cell_value(ws_rozrahunok, f'K{row_num}', ws_formulas)
                    if h_val and k_val:
                        all_plan_items[row_num] = (h_val, k_val)
                        print(f"      - Рядок {row_num}: {h_val} = {k_val}")

        used_cell_formula = ws_formulas[f'{cols["used"]}{zahid["start_row"]}']
        used_cell_value = get_cell_value(ws_rozrahunok, f'{cols["used"]}{zahid["start_row"]}', ws_formulas)

        used_sums = {}
        has_used_data = False

        if used_cell_formula.value and isinstance(used_cell_formula.value, str) and used_cell_formula.value.startswith(
                '='):
            has_used_data = True
            formula = used_cell_formula.value[1:]
            print(f"    Використано {cols['used']}: {used_cell_formula.value} (формула)")

            used_values = set()
            refs = parse_formula_references(used_cell_formula.value)
            for ref in refs:
                row_match = re.search(r'\d+', ref)
                if row_match:
                    row_num = int(row_match.group())
                    k_val = get_cell_value(ws_rozrahunok, f'K{row_num}', ws_formulas)
                    if k_val:
                        used_values.add(k_val)

            numbers = re.findall(r'\b(\d+\.?\d*)\b', formula)
            for num_str in numbers:
                try:
                    num = float(num_str)
                    used_values.add(num)
                except:
                    pass

            print(f"      Використані суми: {used_values}")

            tolerance = 100

            for row_num, (name, plan_suma) in all_plan_items.items():
                closest_used = find_closest_match(plan_suma, used_values, tolerance)

                if closest_used is not None:
                    used_sums[plan_suma] = closest_used
                    diff = plan_suma - closest_used
                    print(f"      ✓ Пара: {plan_suma} (план) ≈ {closest_used} (використано), залишок = {diff:.2f}")
                else:
                    used_sums[plan_suma] = 0
                    print(f"      ⚠️ Не знайдено використання для {plan_suma} (повний залишок)")

        elif used_cell_value and isinstance(used_cell_value, (int, float)) and used_cell_value > 0:
            has_used_data = True
            print(f"    Використано {cols['used']}: {used_cell_value} (число)")

            plan_total = sum(suma for _, suma in all_plan_items.values())
            print(f"      План загалом: {plan_total:.2f}")
            print(f"      Використано: {used_cell_value:.2f}")
            print(f"      Залишок розрахунковий: {plan_total - used_cell_value:.2f}")

            tolerance = 0.1
            from itertools import combinations

            found_combination = False
            for r in range(1, len(all_plan_items) + 1):
                for combo in combinations(all_plan_items.items(), r):
                    combo_sum = sum(suma for _, (_, suma) in combo)
                    if abs(combo_sum - real_zalushok) < tolerance:
                        print(f"      ✓ Знайдено комбінацію товарів (залишок):")
                        for row_num, (name, suma) in combo:
                            print(f"        Рядок {row_num}: {name} = {suma:.2f}")
                            used_sums[suma] = 0
                        found_combination = True
                        break
                if found_combination:
                    break

            if found_combination:
                for row_num, (name, suma) in all_plan_items.items():
                    if suma not in used_sums:
                        used_sums[suma] = suma
                        print(f"      ✗ Рядок {row_num}: {name} = {suma:.2f} (ВИКОРИСТАНО ПОВНІСТЮ)")

        else:
            print(f"    Використано {cols['used']}: порожньо або 0")

        print(f"    Товари що залишились:")

        if has_used_data:
            calculated_total = 0

            for row_num, (name, plan_suma) in all_plan_items.items():
                if plan_suma in used_sums:
                    used_suma = used_sums[plan_suma]
                    ostacha_tovaru = plan_suma - used_suma

                    if abs(ostacha_tovaru) < 0.01:
                        print(f"      ✗ Рядок {row_num}: {name} = {plan_suma:.2f} (ВИКОРИСТАНО ПОВНІСТЮ)")
                    elif ostacha_tovaru > 0.01:
                        print(
                            f"      ✓ Рядок {row_num}: {name} = {plan_suma:.2f} - {used_suma:.2f} = {ostacha_tovaru:.2f}")
                        calculated_total += ostacha_tovaru

                        tovary_by_kekv[kekv].append({
                            'name': name,
                            'suma': ostacha_tovaru,
                            'kekv': kekv,
                            'row': row_num
                        })
                else:
                    print(f"      ✓ Рядок {row_num}: {name} = {plan_suma:.2f} (НЕ ВИКОРИСТАНО)")
                    calculated_total += plan_suma

                    tovary_by_kekv[kekv].append({
                        'name': name,
                        'suma': plan_suma,
                        'kekv': kekv,
                        'row': row_num
                    })

            print(f"\n    ✓ ПЕРЕВІРКА: Розрахований залишок = {calculated_total:.2f}, Реальний = {real_zalushok:.2f}")

            if abs(calculated_total - real_zalushok) < 0.01:
                print(f"      ✅ ЗБІГАЄТЬСЯ!")
            else:
                diff = real_zalushok - calculated_total
                print(f"      ⚠️ РІЗНИЦЯ: {diff:.2f}")

                if abs(diff) > 1:
                    print(f"\n      ⚠️ УВАГА: Розбіжність більше 1 грн!")
                    confirm = input(f"      Продовжити? (1-так, 0-скасувати): ").strip()
                    if confirm != '1':
                        print(f"      ✗ Скасовано користувачем")
                        tovary_by_kekv[kekv] = []
        else:
            print(f"    Шукаємо товар з сумою ≈ {real_zalushok}")
            tolerance = 0.01

            found = False
            for row_num, (name, suma) in all_plan_items.items():
                if abs(suma - real_zalushok) < tolerance:
                    print(f"      ✓ Рядок {row_num}: {name} = {suma} (СПІВПАДАЄ З ЗАЛИШКОМ)")
                    tovary_by_kekv[kekv].append({
                        'name': name,
                        'suma': suma,
                        'kekv': kekv,
                        'row': row_num
                    })
                    found = True
                else:
                    print(f"      ✗ Рядок {row_num}: {name} = {suma}")

            if not found:
                print(f"      ⚠ Не знайдено товар з точною сумою залишку!")

                if len(all_plan_items) == 1:
                    row_num, (name, suma) = list(all_plan_items.items())[0]
                    print(f"\n      🤔 Знайдено ТІЛЬКИ ОДИН товар:")
                    print(f"         Рядок {row_num}: {name} = {suma:.2f}")
                    print(f"         Залишок: {real_zalushok:.2f}")
                    print(f"         Різниця: {abs(suma - real_zalushok):.2f}")

                    confirm = input(
                        f"\n      ❓ Залишок {real_zalushok:.2f} належить товару '{name}'?\n"
                        f"         1 - Так, підтверджую\n"
                        f"         2 - Ні, скасувати\n"
                        f"      👉 Ваш вибір: ").strip()

                    if confirm == '1':
                        print(f"      ✓ Підтверджено: переносимо '{name}'")
                        tovary_by_kekv[kekv].append({
                            'name': name,
                            'suma': real_zalushok,
                            'kekv': kekv,
                            'row': row_num
                        })
                    else:
                        print(f"      ✗ Скасовано користувачем")
                else:
                    print(
                        f"      → Переносимо ВСІ товари (сума = {sum(suma for _, suma in all_plan_items.values()):.2f})")
                    for row_num, (name, suma) in all_plan_items.items():
                        print(f"      ✓ Рядок {row_num}: {name} = {suma} (ПЕРЕНОСИМО)")
                        tovary_by_kekv[kekv].append({
                            'name': name,
                            'suma': suma,
                            'kekv': kekv,
                            'row': row_num
                        })

    wb_formulas.close()
    return tovary_by_kekv


def read_koshtorys_data(ws_koshtorys):
    """Читає дані з Кошторису"""

    event_name = ws_koshtorys['D12'].value
    pp_number = "поточний захід з кошторису"

    print(f"\n📖 Читання Кошторису:")
    print(f"  п/п: {pp_number}")
    print(f"  Назва заходу: {event_name}")

    items = []
    purchased_items = []  # ← ОКРЕМИЙ СПИСОК ДЛЯ КУПЛЕНИХ ТОВАРІВ
    razom_row = None

    print(f"\n📦 Обробка товарів:")

    for row in range(27, ws_koshtorys.max_row + 1):
        c_val = ws_koshtorys[f'C{row}'].value

        if c_val and isinstance(c_val, str) and 'Разом за кошторисом' in c_val:
            razom_row = row
            print(f"  ✓ Знайдено 'Разом за кошторисом' на рядку {row}")
            break

        if c_val and isinstance(c_val, str) and 'Нагородна атрибутика' in c_val:
            print(f"  ⚠️  Рядок {row}: 'Нагородна атрибутика' (пропускаємо)")
            continue

        g_val = ws_koshtorys[f'G{row}'].value
        k_val = ws_koshtorys[f'K{row}'].value

        if not c_val or not g_val or not k_val:
            continue

        kekv = str(g_val).strip()

        suma = k_val
        if isinstance(suma, str):
            try:
                suma = float(suma.replace(',', '.').replace(' ', ''))
            except:
                suma = 0
        else:
            suma = float(suma) if suma else 0

        if suma == 0:
            continue

        item_data = {
            'name': str(c_val).strip(),
            'kekv': kekv,
            'suma': suma,
            'row': row
        }

        # ✅ ПЕРЕВІРЯЄМО ЧИ ЦЕ КУПЛЕНИЙ ТОВАР (1812)
        if is_purchased_item(item_data['name']):
            purchased_items.append(item_data)
            print(f"  [{row}] КЕКВ {kekv}: {c_val} = {suma:.2f} 🛒 (КУПЛЕНО 1812)")
        else:
            items.append(item_data)
            print(f"  [{row}] КЕКВ {kekv}: {c_val} = {suma:.2f}")

    # РОЗРАХОВУЄМО СУМИ ПО КЕКВ
    print(f"\n💰 Розрахунок сум по КЕКВ:")

    kekv_data = {}  # Загальна сума (з купленими)
    kekv_purchased = {}  # Тільки куплені товари (1812)

    # Звичайні товари
    for item in items:
        kekv = item['kekv']
        if kekv not in kekv_data:
            kekv_data[kekv] = 0
        kekv_data[kekv] += item['suma']

    # Куплені товари (1812) - додаємо до загальної суми та окремо враховуємо
    for item in purchased_items:
        kekv = item['kekv']
        if kekv not in kekv_data:
            kekv_data[kekv] = 0
        kekv_data[kekv] += item['suma']

        if kekv not in kekv_purchased:
            kekv_purchased[kekv] = 0
        kekv_purchased[kekv] += item['suma']

    for kekv in sorted(kekv_data.keys()):
        suma = kekv_data[kekv]
        purchased = kekv_purchased.get(kekv, 0)
        unpurchased = suma - purchased
        if purchased > 0:
            print(f"  КЕКВ {kekv}: {suma:.2f} грн")
            print(f"    ├─ не куплено: {unpurchased:.2f}")
            print(f"    └─ куплено (1812): {purchased:.2f}")
        else:
            print(f"  КЕКВ {kekv}: {suma:.2f} грн")

    # ПЕРЕВІРКА
    if razom_row:
        razom_suma = ws_koshtorys[f'K{razom_row}'].value or 0
        if isinstance(razom_suma, str):
            try:
                razom_suma = float(razom_suma.replace(',', '.').replace(' ', ''))
            except:
                razom_suma = 0

        total_tovarov = sum(kekv_data.values())

        print(f"\n✓ Перевірка:")
        print(f"  'Разом за кошторисом' (K{razom_row}): {razom_suma:.2f}")
        print(f"  Сума по КЕКВ: {total_tovarov:.2f}")

        diff = abs(razom_suma - total_tovarov)
        if diff < 0.01:
            print(f"  Співпадає: ✓")
        else:
            print(f"  ⚠️  Різниця: {diff:.2f} грн")

    return pp_number, event_name, items, kekv_data, kekv_purchased, purchased_items


def find_zahid_by_name_in_rozrahunok(ws_rozrahunok, event_name):
    """Знаходить захід в Розрахунку за назвою"""
    if not event_name:
        return None

    event_name_normalized = str(event_name).strip().lower()
    event_name_normalized = event_name_normalized.strip('"').strip("'").strip('«').strip('»').strip()
    event_name_normalized = event_name_normalized.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
    event_name_normalized = ' '.join(event_name_normalized.split())

    for row in range(2, ws_rozrahunok.max_row + 1):
        b_val = ws_rozrahunok[f'B{row}'].value
        if b_val:
            b_val_normalized = str(b_val).strip().lower()
            b_val_normalized = b_val_normalized.strip('"').strip("'").strip('«').strip('»').strip()
            b_val_normalized = b_val_normalized.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
            b_val_normalized = ' '.join(b_val_normalized.split())

            if event_name_normalized == b_val_normalized:
                a_val = ws_rozrahunok[f'A{row}'].value
                c_val = ws_rozrahunok[f'C{row}'].value

                if a_val is not None and isinstance(a_val, (int, float)):
                    return {
                        'number': int(a_val),
                        'termin': c_val
                    }

    event_words = set(word for word in event_name_normalized.split() if len(word) >= 3)
    best_match = None
    max_common_words = 0

    for row in range(2, ws_rozrahunok.max_row + 1):
        b_val = ws_rozrahunok[f'B{row}'].value
        if b_val:
            b_val_normalized = str(b_val).strip().lower()
            b_val_normalized = b_val_normalized.strip('"').strip("'").strip('«').strip('»').strip()
            b_val_normalized = b_val_normalized.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
            b_val_normalized = ' '.join(b_val_normalized.split())

            b_words = set(word for word in b_val_normalized.split() if len(word) >= 3)
            common_words = event_words & b_words
            num_common = len(common_words)

            if num_common > max_common_words and num_common >= 3:
                max_common_words = num_common
                best_match = {
                    'row': row,
                    'name': b_val_normalized,
                    'common_words': common_words
                }

    if best_match:
        row = best_match['row']
        a_val = ws_rozrahunok[f'A{row}'].value
        c_val = ws_rozrahunok[f'C{row}'].value

        if a_val is not None and isinstance(a_val, (int, float)):
            return {
                'number': int(a_val),
                'termin': c_val
            }

    return None


def insert_data_to_zminy(zminy_file, koshtorys_data, zahody_rozrahunok, ostacha_items, koshtorys_file, rozrahunok_file):
    """Вставляє дані в Зміни.xlsx"""

    _, event_name, items_koshtorys, kekv_data, kekv_purchased, purchased_items = koshtorys_data

    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    times_font = Font(name='Times New Roman', size=11)
    times_font_bold = Font(name='Times New Roman', size=11, bold=True)

    if zminy_file.exists():
        wb_zminy = openpyxl.load_workbook(zminy_file)
        ws_zminy = wb_zminy.active
    else:
        wb_zminy = openpyxl.Workbook()
        ws_zminy = wb_zminy.active
        # Створюємо заголовки
        headers = ['п/п', 'Назва заходу', 'Термін, місце проведення', 'Найменування',
                   'Сума', 'КЕКВ 2210', 'КЕКВ 2240', 'КЕКВ 2250', 'Сума витрат']

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col, header in enumerate(headers, start=1):
            cell = ws_zminy.cell(row=1, column=col, value=header)
            cell.font = times_font
            cell.alignment = center_alignment
            cell.border = border_style

        ws_zminy.column_dimensions['A'].width = 10
        ws_zminy.column_dimensions['B'].width = 20
        ws_zminy.column_dimensions['C'].width = 20
        ws_zminy.column_dimensions['D'].width = 23
        ws_zminy.column_dimensions['E'].width = 12
        ws_zminy.column_dimensions['F'].width = 15
        ws_zminy.column_dimensions['G'].width = 15
        ws_zminy.column_dimensions['H'].width = 15
        ws_zminy.column_dimensions['I'].width = 15

    # Зберігаємо старі дані
    old_data = []
    for row in range(2, ws_zminy.max_row + 1):
        row_data = []
        has_data = False
        for col in range(1, 10):
            val = ws_zminy.cell(row=row, column=col).value
            row_data.append(val)
            if val is not None:
                has_data = True
        if has_data:
            old_data.append(row_data)

    # Старі блоки могли мати об'єднані комірки. Перед очищенням їх треба роз'єднати,
    # інакше openpyxl не дозволить змінювати MergedCell.
    for merged_range in list(ws_zminy.merged_cells.ranges):
        if merged_range.max_row >= 2:
            ws_zminy.unmerge_cells(str(merged_range))

    # Очищаємо
    for row in ws_zminy.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()
            cell.border = Border()

    current_row = 2
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    number_format = '#,##0.00'

    # ========== ЗАХОДИ З РОЗРАХУНКУ (З МІНУСОМ!) ==========

    for zahid_data in zahody_rozrahunok:
        zahid = zahid_data['zahid']
        tovary = zahid_data['tovary']
        zalushky = zahid_data.get('zalushky', {})

        kekv_sums = {
            '2210': -zalushky.get('2210', 0) if tovary['2210'] else 0,
            '2240': -zalushky.get('2240', 0) if tovary['2240'] else 0,
            '2250': -zalushky.get('2250', 0) if tovary['2250'] else 0
        }
        total_zahid = sum(kekv_sums.values())

        if total_zahid == 0:
            continue

        zahid_start_row = current_row

        ws_zminy.cell(row=current_row, column=1, value=zahid['number'])
        ws_zminy.cell(row=current_row, column=2, value=zahid['name'])
        ws_zminy.cell(row=current_row, column=3, value=zahid['termin'])

        ws_zminy.cell(row=current_row, column=6, value=kekv_sums['2210'])
        ws_zminy.cell(row=current_row, column=7, value=kekv_sums['2240'])
        ws_zminy.cell(row=current_row, column=8, value=kekv_sums['2250'])
        ws_zminy.cell(row=current_row, column=9, value=total_zahid)

        for col in range(1, 10):
            cell = ws_zminy.cell(row=current_row, column=col)
            cell.alignment = center_alignment
            cell.border = border_style
            cell.font = times_font
            if col >= 6:
                cell.number_format = number_format

        first_added = False
        for kekv in ['2210', '2240', '2250']:
            if tovary[kekv] and not first_added:
                first_tovar = tovary[kekv][0]
                ws_zminy.cell(row=current_row, column=4, value=first_tovar['name'])
                ws_zminy.cell(row=current_row, column=5, value=-first_tovar['suma'])

                ws_zminy.cell(row=current_row, column=4).font = times_font
                ws_zminy.cell(row=current_row, column=5).font = times_font
                ws_zminy.cell(row=current_row, column=5).number_format = number_format

                tovary[kekv].pop(0)
                first_added = True
                break

        current_row += 1

        for kekv in ['2210', '2240', '2250']:
            if tovary[kekv]:
                for tovar in tovary[kekv]:
                    ws_zminy.cell(row=current_row, column=4, value=tovar['name'])
                    ws_zminy.cell(row=current_row, column=5, value=-tovar['suma'])

                    for col in range(1, 10):
                        cell = ws_zminy.cell(row=current_row, column=col)
                        cell.alignment = center_alignment
                        cell.border = border_style
                        cell.font = times_font

                    ws_zminy.cell(row=current_row, column=5).number_format = number_format
                    current_row += 1

        zahid_end_row = current_row - 1

        if zahid_end_row > zahid_start_row:
            merge_columns = [1, 2, 3, 6, 7, 8, 9]

            for col in merge_columns:
                ws_zminy.merge_cells(
                    start_row=zahid_start_row,
                    start_column=col,
                    end_row=zahid_end_row,
                    end_column=col
                )
                for row in range(zahid_start_row, zahid_end_row + 1):
                    ws_zminy.cell(row=row, column=col).border = border_style

    # ========== ЗАХІД З КОШТОРИСУ (БЕЗ МІНУСА) ==========

    koshtorys_start_row = current_row

    # Використовуємо переданий rozrahunok_file — без жодного хардкоду назви!
    wb_rozrahunok_for_search = openpyxl.load_workbook(rozrahunok_file, data_only=True)
    ws_rozrahunok_for_search = wb_rozrahunok_for_search.active

    zahid_info = find_zahid_by_name_in_rozrahunok(ws_rozrahunok_for_search, event_name)
    wb_rozrahunok_for_search.close()

    pp_number = None
    termin = None

    if zahid_info:
        pp_number = zahid_info['number']
        termin = zahid_info['termin']
    else:
        print(f"\n⚠️  Захід '{event_name}' не знайдено в Розрахунку")
        print(f"   Стовпці A та C залишаться порожніми")

    ws_zminy.cell(row=current_row, column=1, value=pp_number)
    ws_zminy.cell(row=current_row, column=2, value=event_name)
    ws_zminy.cell(row=current_row, column=3, value=termin)

    # РОЗРАХУНОК СУМ ПО КЕКВ (БЕЗ куплених + остача + куплені)
    kekv_sums_for_display = {
        '2210': (kekv_data.get('2210', 0) - kekv_purchased.get('2210', 0)) + ostacha_items.get('2210',
                                                                                               0) + kekv_purchased.get(
            '2210', 0),
        '2240': (kekv_data.get('2240', 0) - kekv_purchased.get('2240', 0)) + ostacha_items.get('2240',
                                                                                               0) + kekv_purchased.get(
            '2240', 0),
        '2250': (kekv_data.get('2250', 0) - kekv_purchased.get('2250', 0)) + ostacha_items.get('2250',
                                                                                               0) + kekv_purchased.get(
            '2250', 0)
    }

    total_sum = sum(kekv_sums_for_display.values())

    ws_zminy.cell(row=current_row, column=6, value=kekv_sums_for_display['2210'])
    ws_zminy.cell(row=current_row, column=7, value=kekv_sums_for_display['2240'])
    ws_zminy.cell(row=current_row, column=8, value=kekv_sums_for_display['2250'])
    ws_zminy.cell(row=current_row, column=9, value=total_sum)

    for col in range(1, 10):
        cell = ws_zminy.cell(row=current_row, column=col)
        cell.alignment = center_alignment
        cell.border = border_style
        cell.font = times_font
        if col >= 6:
            cell.number_format = number_format

    # Перший товар з Кошторису
    if items_koshtorys:
        first_item = items_koshtorys[0]
        ws_zminy.cell(row=current_row, column=4, value=first_item['name'])
        ws_zminy.cell(row=current_row, column=5, value=first_item['suma'])

        ws_zminy.cell(row=current_row, column=4).font = times_font
        ws_zminy.cell(row=current_row, column=5).font = times_font
        ws_zminy.cell(row=current_row, column=5).number_format = number_format

        current_row += 1

        # Решта товарів
        for item in items_koshtorys[1:]:
            ws_zminy.cell(row=current_row, column=4, value=item['name'])
            ws_zminy.cell(row=current_row, column=5, value=item['suma'])

            for col in range(1, 10):
                cell = ws_zminy.cell(row=current_row, column=col)
                cell.alignment = center_alignment
                cell.border = border_style
                cell.font = times_font

            ws_zminy.cell(row=current_row, column=5).number_format = number_format
            current_row += 1

    # РЯДКИ ОСТАЧІ (БЛАКИТНІ)
    for kekv, ostacha_suma in ostacha_items.items():
        if ostacha_suma > 0:
            ws_zminy.cell(row=current_row, column=4, value=OSTACHA_NAMES[kekv])
            ws_zminy.cell(row=current_row, column=5, value=ostacha_suma)

            for col in range(1, 10):
                cell = ws_zminy.cell(row=current_row, column=col)
                cell.alignment = center_alignment
                cell.border = border_style
                cell.font = times_font

            ws_zminy.cell(row=current_row, column=4).fill = OSTACHA_COLOR
            ws_zminy.cell(row=current_row, column=5).fill = OSTACHA_COLOR
            ws_zminy.cell(row=current_row, column=5).number_format = number_format

            current_row += 1

    # ДОДАЄМО КУПЛЕНІ ТОВАРИ (1812) - ЗВИЧАЙНІ РЯДКИ
    print(f"\n💳 Додавання куплених товарів (1812):")

    for item in purchased_items:
        ws_zminy.cell(row=current_row, column=4, value=item['name'])
        ws_zminy.cell(row=current_row, column=5, value=item['suma'])

        for col in range(1, 10):
            cell = ws_zminy.cell(row=current_row, column=col)
            cell.alignment = center_alignment
            cell.border = border_style
            cell.font = times_font

        ws_zminy.cell(row=current_row, column=5).number_format = number_format

        print(f"  ✓ Додано: {item['name']} - {item['suma']:.2f} грн (КЕКВ {item['kekv']})")
        current_row += 1

    koshtorys_end_row = current_row - 1

    # ОБ'ЄДНАННЯ для заходу з Кошторису
    if koshtorys_end_row > koshtorys_start_row:
        merge_columns = [1, 2, 3, 6, 7, 8, 9]

        for col in merge_columns:
            ws_zminy.merge_cells(
                start_row=koshtorys_start_row,
                start_column=col,
                end_row=koshtorys_end_row,
                end_column=col
            )
            for row in range(koshtorys_start_row, koshtorys_end_row + 1):
                ws_zminy.cell(row=row, column=col).border = border_style

    # Повертаємо старі дані
    for old_row_data in old_data:
        for col, value in enumerate(old_row_data, start=1):
            cell = ws_zminy.cell(row=current_row, column=col, value=value)
            cell.border = border_style
            cell.font = times_font
            cell.alignment = center_alignment
            if col in [5, 6, 7, 8, 9] and isinstance(value, (int, float)):
                cell.number_format = number_format
        current_row += 1

    wb_zminy.save(zminy_file)
    wb_zminy.close()

    print(f"\n✅ Дані успішно записано в {zminy_file}")
    print(f"   Всього рядків: {current_row - 1}")


def choose_kekv_for_zahid(zahid, zalushky, koshtorys_needs):
    """Діалог вибору КЕКВ для конкретного заходу"""

    print("\n" + "=" * 80)
    print(f"📋 ЗАХІД #{zahid['number']}: {zahid['name']}")
    print("=" * 80)

    print(f"\n💰 ЗАЛИШКИ на цьому заході:")
    total_zalushky = 0
    for kekv in ['2210', '2240', '2250']:
        suma = zalushky.get(kekv, 0)
        if suma > 0:
            print(f"   КЕКВ {kekv}: {suma:>10.2f} грн")
            total_zalushky += suma

    if total_zalushky == 0:
        print(f"   ⚠️  Немає залишків на цьому заході!")
        return []

    print(f"   {'─' * 30}")
    print(f"   РАЗОМ:      {total_zalushky:>10.2f} грн")

    print(f"\n📊 ПОТРЕБИ згідно Кошторису (БЕЗ куплених товарів):")
    total_needs = 0
    for kekv in ['2210', '2240', '2250']:
        suma = koshtorys_needs.get(kekv, 0)
        if suma > 0:
            print(f"   КЕКВ {kekv}: {suma:>10.2f} грн")
            total_needs += suma

    if total_needs > 0:
        print(f"   {'─' * 30}")
        print(f"   РАЗОМ:      {total_needs:>10.2f} грн")

    if total_zalushky > 0 and total_needs > 0:
        diff = total_zalushky - total_needs
        per_kekv_deficits = {
            kekv: koshtorys_needs.get(kekv, 0) - zalushky.get(kekv, 0)
            for kekv in ['2210', '2240', '2250']
            if koshtorys_needs.get(kekv, 0) > zalushky.get(kekv, 0)
        }
        print(f"\n⚖️  БАЛАНС:")
        if per_kekv_deficits:
            print_big_deficit_warning(per_kekv_deficits)
            if diff >= 0:
                print(f"   Загалом гроші є (+{diff:.2f} грн), але по окремих КЕКВ НЕ ВИСТАЧАЄ.")
            else:
                print(f"   Загалом теж не вистачає: {abs(diff):.2f} грн")
        elif diff >= 0:
            print(f"   Залишків достатньо! Лишиться: {diff:.2f} грн")
        else:
            print(f"   ⚠️  Не вистачає: {abs(diff):.2f} грн")

    print(f"\n🎯 З яких КЕКВ забирати залишки?")
    print(f"   Доступні опції:")

    available_kekv = []
    option_map = {}
    option_num = 1

    for kekv in ['2210', '2240', '2250']:
        if zalushky.get(kekv, 0) > 0:
            print(f"   {option_num} - Тільки КЕКВ {kekv} ({zalushky[kekv]:.2f} грн)")
            option_map[str(option_num)] = [kekv]
            available_kekv.append(kekv)
            option_num += 1

    if len(available_kekv) >= 2:
        for i in range(len(available_kekv)):
            for j in range(i + 1, len(available_kekv)):
                kekv1, kekv2 = available_kekv[i], available_kekv[j]
                suma = zalushky[kekv1] + zalushky[kekv2]
                print(f"   {option_num} - КЕКВ {kekv1} + {kekv2} ({suma:.2f} грн)")
                option_map[str(option_num)] = [kekv1, kekv2]
                option_num += 1

    if len(available_kekv) >= 2:
        print(f"   {option_num} - Всі КЕКВ ({total_zalushky:.2f} грн)")
        option_map[str(option_num)] = available_kekv
        option_num += 1

    print(f"   0 - Пропустити цей захід (не забирати залишки)")

    while True:
        choice = input(f"\n👉 Ваш вибір (0-{option_num - 1}): ").strip()

        if choice == '0':
            print(f"   ⏭️  Захід #{zahid['number']} пропущено")
            return []

        if choice in option_map:
            selected_kekv = option_map[choice]
            selected_suma = sum(zalushky[k] for k in selected_kekv)

            print(f"\n   ✓ Обрано: {', '.join(selected_kekv)}")
            print(f"   💵 Сума: {selected_suma:.2f} грн")

            return selected_kekv
        else:
            print(f"   ❌ Некоректний вибір! Спробуйте ще раз.")


def choose_take_mode_for_zahid(zahid, selected_kekv, zalushky, koshtorys_needs, already_taken):
    """Питає, скільки грошей брати з обраних КЕКВ."""
    remaining_needs = get_remaining_needs(koshtorys_needs, already_taken)
    selected_available = sum(zalushky.get(kekv, 0) for kekv in selected_kekv)
    selected_needed = sum(remaining_needs.get(kekv, 0) for kekv in selected_kekv)

    print("\n" + "-" * 80)
    print(f"💵 СКІЛЬКИ ГРОШЕЙ ВЗЯТИ З ЗАХОДУ #{zahid['number']}?")
    print("-" * 80)
    print("Обрані КЕКВ:")

    deficits = {}
    for kekv in selected_kekv:
        available = zalushky.get(kekv, 0)
        needed = remaining_needs.get(kekv, 0)
        diff = available - needed
        print(f"   КЕКВ {kekv}: доступно {available:.2f} грн, ще потрібно {needed:.2f} грн")
        if needed > 0 and diff < 0:
            deficits[kekv] = abs(diff)

    if deficits:
        print_big_deficit_warning(deficits)

    if selected_needed <= 0:
        print("\n⚠️  За обраними КЕКВ потреба вже закрита.")
        print("   Режим 'тільки скільки потрібно' нічого не забере.")

    print(f"\nДоступно з обраних КЕКВ: {selected_available:.2f} грн")
    print(f"Ще потрібно по обраних КЕКВ: {selected_needed:.2f} грн")

    print("\nОберіть режим:")
    print("   1 - Взяти ТІЛЬКИ скільки потрібно за Кошторисом")
    print("       (зайві гроші НЕ підуть у сині рядки остачі)")
    print("   2 - Взяти ВСЕ з обраних КЕКВ")
    print("       (зайві гроші підуть у сині рядки остачі)")
    print("   0 - Пропустити цей захід")

    while True:
        choice = input("\n👉 Ваш вибір режиму (0-2): ").strip()

        if choice == '0':
            print(f"   ⏭️  Захід #{zahid['number']} пропущено")
            return None
        if choice == '1':
            print("   ✓ Режим: беремо тільки потрібну суму")
            return 'needed'
        if choice == '2':
            print("   ✓ Режим: беремо все, надлишок піде в остачу")
            return 'all'

        print("   ❌ Некоректний вибір! Спробуйте ще раз.")


def calculate_take_amounts(selected_kekv, zalushky, koshtorys_needs, already_taken, take_mode):
    """Повертає суми, які реально забираємо по КЕКВ."""
    take_amounts = {}
    remaining_needs = get_remaining_needs(koshtorys_needs, already_taken)

    for kekv in selected_kekv:
        available = zalushky.get(kekv, 0)
        if take_mode == 'all':
            amount = available
        else:
            amount = min(available, remaining_needs.get(kekv, 0))

        if amount > 0:
            take_amounts[kekv] = amount

    return take_amounts


def limit_tovary_to_take_amounts(tovary_by_kekv, take_amounts):
    """Обрізає список товарів так, щоб їх сума дорівнювала сумі, яку забираємо."""
    limited = {'2210': [], '2240': [], '2250': []}

    for kekv in ['2210', '2240', '2250']:
        target = take_amounts.get(kekv, 0)
        remaining = target

        if remaining <= 0:
            continue

        for item in tovary_by_kekv.get(kekv, []):
            if remaining <= 0.01:
                break

            item_sum = item.get('suma', 0)
            if item_sum <= 0:
                continue

            new_item = item.copy()
            if item_sum > remaining:
                new_item['suma'] = remaining
                limited[kekv].append(new_item)
                remaining = 0
                break

            limited[kekv].append(new_item)
            remaining -= item_sum

        if remaining > 0.01:
            limited[kekv].append({
                'name': f"Сума до переносу КЕКВ {kekv}",
                'suma': remaining,
                'kekv': kekv,
                'row': None
            })

    return limited


def main():
    """Головна функція програми"""

    # Шляхи до файлів за замовчуванням
    default_rozrahunok = 'Розрахунок до календарного плану 2025а (Автозбережено).xlsx'
    default_koshtorys = 'Кошторис.xlsx'

    rozrahunok_file = find_rozrahunok_file(default_rozrahunok)
    koshtorys_file = Path(default_koshtorys)
    zminy_file = Path('Зміни.xlsx')

    print("=" * 80)
    print("🚀 ПРОГРАМА ПЕРЕНОСУ ДАНИХ З РОЗРАХУНКУ ТА КОШТОРИСУ В ЗМІНИ")
    print("=" * 80)

    # ========== ПЕРЕВІРКА ФАЙЛУ РОЗРАХУНОК ==========
    if not rozrahunok_file:
        print(f"\n⚠️  Файл '{default_rozrahunok}' не знайдено!")
        print(f"📂 Оберіть файл Розрахунку вручну...")

        selected_file = select_file(
            "Оберіть файл Розрахунку",
            default_rozrahunok
        )

        if selected_file:
            rozrahunok_file = selected_file
            print(f"✅ Обрано файл: {rozrahunok_file.name}")
        else:
            print(f"\n❌ ПОМИЛКА: Файл Розрахунку не обрано! Програма завершена.")
            return
    else:
        print(f"✅ Знайдено файл Розрахунку: {rozrahunok_file.name}")

    # ========== ПЕРЕВІРКА ФАЙЛУ КОШТОРИС ==========
    if not koshtorys_file.exists():
        print(f"\n⚠️  Файл '{default_koshtorys}' не знайдено!")
        print(f"📂 Оберіть файл Кошторису вручну...")

        selected_file = select_file(
            "Оберіть файл Кошторису",
            default_koshtorys
        )

        if selected_file:
            koshtorys_file = selected_file
            print(f"✅ Обрано файл: {koshtorys_file.name}")
        else:
            print(f"\n❌ ПОМИЛКА: Файл Кошторису не обрано! Програма завершена.")
            return
    else:
        print(f"✅ Знайдено файл Кошторису: {koshtorys_file.name}")

    # КРОК 1: Запитуємо номери заходів з Розрахунку
    print("\n" + "=" * 80)
    print("КРОК 1: Введення номерів заходів з Розрахунку")
    print("=" * 80)

    zahid_numbers_input = input("\nВведіть номери заходів через кому (наприклад: 1,3,5,9): ").strip()

    if not zahid_numbers_input:
        print("\n⚠️  Не введено жодного номера заходу!")
        return

    try:
        zahid_numbers = [int(x.strip()) for x in zahid_numbers_input.split(',')]
        print(f"✓ Будемо обробляти заходи: {zahid_numbers}")
    except ValueError:
        print("\n❌ ПОМИЛКА: Введено некоректні номери заходів!")
        return

    # КРОК 2: Читаємо Кошторис (щоб знати потреби)
    print("\n" + "=" * 80)
    print("КРОК 2: Аналіз Кошторису (визначення потреб)")
    print("=" * 80)

    wb_koshtorys = openpyxl.load_workbook(koshtorys_file, data_only=True)
    ws_koshtorys = wb_koshtorys.active

    koshtorys_data = read_koshtorys_data(ws_koshtorys)
    koshtorys_needs = get_koshtorys_needs(koshtorys_data)

    wb_koshtorys.close()

    print(f"\n📊 Загальні потреби з Кошторису:")
    total_need = 0
    for kekv in ['2210', '2240', '2250']:
        if koshtorys_needs[kekv] > 0:
            print(f"   КЕКВ {kekv}: {koshtorys_needs[kekv]:.2f} грн")
            total_need += koshtorys_needs[kekv]
    print(f"   РАЗОМ: {total_need:.2f} грн")

    # КРОК 3: Обробка кожного заходу окремо
    print("\n" + "=" * 80)
    print("КРОК 3: Вибір КЕКВ для кожного заходу")
    print("=" * 80)

    wb_rozrahunok = openpyxl.load_workbook(rozrahunok_file, data_only=True)
    wb_rozrahunok_formulas = openpyxl.load_workbook(rozrahunok_file, data_only=False)
    ws_rozrahunok = wb_rozrahunok.active
    ws_rozrahunok_formulas = wb_rozrahunok_formulas.active

    zahody_rozrahunok = []
    total_ostacha = {'2210': 0, '2240': 0, '2250': 0}

    for zahid_num in zahid_numbers:
        # Знаходимо захід
        zahid = find_zahid_in_rozrahunok(ws_rozrahunok, zahid_num)

        if not zahid:
            print(f"\n⚠️  Захід #{zahid_num} не знайдено в Розрахунку!")
            continue

        # Отримуємо залишки
        zalushky = get_zalushky_for_zahid(ws_rozrahunok, zahid['start_row'], ws_rozrahunok_formulas)

        # Діалог вибору КЕКВ для цього заходу
        selected_kekv = choose_kekv_for_zahid(zahid, zalushky, koshtorys_needs)

        if not selected_kekv:
            continue  # Користувач пропустив цей захід

        take_mode = choose_take_mode_for_zahid(
            zahid,
            selected_kekv,
            zalushky,
            koshtorys_needs,
            total_ostacha
        )

        if not take_mode:
            continue

        take_amounts = calculate_take_amounts(
            selected_kekv,
            zalushky,
            koshtorys_needs,
            total_ostacha,
            take_mode
        )

        if not take_amounts:
            print(f"\n⚠️  По обраних КЕКВ немає суми для переносу. Захід #{zahid_num} пропущено.")
            continue

        print(f"\n✅ Буде забрано з заходу #{zahid_num}:")
        for kekv in ['2210', '2240', '2250']:
            if take_amounts.get(kekv, 0) > 0:
                print(f"   КЕКВ {kekv}: {take_amounts[kekv]:.2f} грн")

        # Отримуємо товари тільки для обраних КЕКВ
        print(f"\n🔄 Обробка товарів заходу #{zahid_num}...")

        tovary = get_tovary_by_kekv(
            ws_rozrahunok,
            zahid,
            zalushky,
            {},
            selected_kekv,  # Передаємо тільки обрані КЕКВ!
            rozrahunok_file
        )
        tovary = limit_tovary_to_take_amounts(tovary, take_amounts)

        # Додаємо до загальної остачі
        for kekv, amount in take_amounts.items():
            total_ostacha[kekv] += amount

        zahody_rozrahunok.append({
            'zahid': zahid,
            'tovary': tovary,
            'zalushky': take_amounts,  # Зберігаємо тільки суму, яку реально забираємо!
            'take_mode': take_mode
        })

    wb_rozrahunok.close()
    wb_rozrahunok_formulas.close()

    # КРОК 4: Підсумки
    print("\n" + "=" * 80)
    print("КРОК 4: ПІДСУМОК")
    print("=" * 80)

    print(f"\n✅ Оброблено заходів: {len(zahody_rozrahunok)}")

    if len(zahody_rozrahunok) == 0:
        print(f"\n⚠️  Не було обрано жодного заходу для переносу!")
        return

    print(f"\n💰 ЗАГАЛЬНА ОСТАЧА по всіх заходах:")
    total_ostacha_suma = 0
    for kekv in ['2210', '2240', '2250']:
        if total_ostacha[kekv] > 0:
            print(f"   КЕКВ {kekv}: {total_ostacha[kekv]:.2f} грн")
            total_ostacha_suma += total_ostacha[kekv]

    print(f"   {'─' * 30}")
    print(f"   РАЗОМ:      {total_ostacha_suma:.2f} грн")

    print(f"\n📊 ПОТРЕБИ з Кошторису:")
    for kekv in ['2210', '2240', '2250']:
        if koshtorys_needs[kekv] > 0:
            print(f"   КЕКВ {kekv}: {koshtorys_needs[kekv]:.2f} грн")
    print(f"   {'─' * 30}")
    print(f"   РАЗОМ:      {total_need:.2f} грн")

    # БАЛАНС ПО КОЖНОМУ КЕКВ
    print(f"\n⚖️  БАЛАНС ПО КОЖНОМУ КЕКВ:")
    total_deficit = 0
    total_surplus = 0
    deficit_by_kekv = {}

    for kekv in ['2210', '2240', '2250']:
        ostacha = total_ostacha.get(kekv, 0)
        potreba = koshtorys_needs.get(kekv, 0)
        diff = ostacha - potreba

        if ostacha > 0 or potreba > 0:
            if diff >= 0:
                print(f"   КЕКВ {kekv}: {ostacha:.2f} - {potreba:.2f} = +{diff:.2f} (надлишок)")
                total_surplus += diff
            else:
                print(f"   КЕКВ {kekv}: {ostacha:.2f} - {potreba:.2f} = {diff:.2f} (не вистачає)")
                total_deficit += abs(diff)
                deficit_by_kekv[kekv] = abs(diff)

    print(f"   {'─' * 50}")

    overall_diff = total_ostacha_suma - total_need
    if total_deficit > 0:
        if overall_diff >= 0:
            print(f"   ⚠️  ЗАГАЛЬНИЙ БАЛАНС: +{overall_diff:.2f} грн, але по окремих КЕКВ НЕ ВИСТАЧАЄ")
        else:
            print(f"   ⚠️  ЗАГАЛЬНИЙ БАЛАНС: {overall_diff:.2f} грн (не вистачає)")
    elif overall_diff >= 0:
        print(f"   ✅ ЗАГАЛЬНИЙ БАЛАНС: +{overall_diff:.2f} грн (надлишок)")
    else:
        print(f"   ⚠️  ЗАГАЛЬНИЙ БАЛАНС: {overall_diff:.2f} грн (не вистачає)")

    if total_deficit > 0 and total_surplus > 0:
        print(f"\n   💡 Примітка: Є надлишки по одних КЕКВ та дефіцит по інших")

    if total_deficit > 0:
        print_big_deficit_warning(deficit_by_kekv)

        print("\nЩоб продовжити і все одно створити Зміни.xlsx з нестачею коштів,")
        print("потрібно явно підтвердити це рішення.")
        print("   1 - Я БАЧУ, ЩО ГРОШЕЙ НЕ ВИСТАЧАЄ, але все одно створити Зміни")
        print("   0 - Скасувати, Зміни НЕ створювати")

        confirm_deficit = input("\n👉 Ваш вибір (1 або 0): ").strip()
        if confirm_deficit != '1':
            print("\n❌ Створення Зміни.xlsx скасовано через нестачу коштів.")
            return

    # КРОК 5: Розрахунок РЕАЛЬНОЇ остачі
    print("\n" + "=" * 80)
    print("КРОК 5: Розрахунок реальної остачі")
    print("=" * 80)

    # РЕАЛЬНА ОСТАЧА = Забрали з Розрахунку - Потреба з Кошторису
    real_ostacha = {}
    print(f"\n💎 РЕАЛЬНА ОСТАЧА (для синіх рядків):")

    for kekv in ['2210', '2240', '2250']:
        zabrano = total_ostacha.get(kekv, 0)  # Скільки забрали
        potreba = koshtorys_needs.get(kekv, 0)  # Скільки треба
        ostacha = zabrano - potreba  # Реальна остача

        if ostacha > 0:
            real_ostacha[kekv] = ostacha
            print(f"   КЕКВ {kekv}: {zabrano:.2f} - {potreba:.2f} = {ostacha:.2f} грн")
        else:
            real_ostacha[kekv] = 0

    total_real_ostacha = sum(real_ostacha.values())
    if total_real_ostacha > 0:
        print(f"   {'─' * 50}")
        print(f"   РАЗОМ ОСТАЧА: {total_real_ostacha:.2f} грн")
    else:
        print(f"   (Немає остачі - все використано)")

    # КРОК 6: Запис в Зміни
    print("\n" + "=" * 80)
    print("КРОК 6: Запис даних у файл Зміни.xlsx")
    print("=" * 80)

    insert_data_to_zminy(
        zminy_file,
        koshtorys_data,
        zahody_rozrahunok,
        real_ostacha,
        koshtorys_file,
        rozrahunok_file   # ← Передаємо реальний шлях, обраний користувачем!
    )

    print("\n" + "=" * 80)
    print("🎉 ПРОГРАМА ЗАВЕРШЕНА УСПІШНО!")
    print("=" * 80)
    print(f"\n📁 Перевірте файл '{zminy_file}'")
    print(f"\n💡 Рядки з остачею виділено блакитним кольором")


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print(f"\n\n⚠️  Програму перервано користувачем")
    except Exception as e:
        print(f"\n❌ КРИТИЧНА ПОМИЛКА: {e}")
        import traceback

        traceback.print_exc()
    finally:
        try:
            input("\nНатисніть Enter для закриття...")
        except EOFError:
            pass
